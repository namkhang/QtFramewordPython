import glob
import os
import datetime
from sys import path
from Ecercise5UI import *
from PyQt5 import QtCore , QtGui , QtWidgets
from openpyxl import Workbook , load_workbook
class Ecercise5(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.ui.btn_find.pressed.connect(self.Find)
        self.ui.btn_export_excel.pressed.connect(self.Export)
        self.ui.btn_import_excel.pressed.connect(self.Import)
        self.dictResult = dict()  
        self.nameFileExcel = str() 
    def Find(self):
        self.dictResult = dict() #reset lại dictResult
        for i in range(self.ui.tableWidget.rowCount() -1 ,-1 , -1): # vì xóa theo tuần tự từ row cuối lên row đầu tiên nên cho vòng lặp chạy ngược
                self.ui.tableWidget.removeRow(i)  #reset lại bảng hiển thị trước mỗi lần tìm kiếm
        self.ui.label_error.setText('')
        ext = self.ui.text_find_ext.toPlainText()
        path = self.ui.text_find_path.toPlainText()
        timeFind =  self.ui.text_find_time.toPlainText()
        arrResult = glob.glob('*{}'.format(ext))
        if len(arrResult) > 0 : # nếu có file thì chạy hàm xử lý còn không sẽ xuất ra thông báo
            for i in arrResult:
                timeModified = os.path.getmtime(i) #lay ra thời gian lần cuối modified
                timeToNow = datetime.datetime.fromtimestamp(timeModified) 
                time = timeToNow.strftime('%d/%m/%Y') 
                print(time)
                if time == timeFind:
                    self.dictResult[i] = path
            if bool(self.dictResult) == True:
                row = 0 
                for key in self.dictResult :
                    self.ui.tableWidget.insertRow(row) # tạo stt của row
                    self.ui.tableWidget.setItem(row , 0,QtWidgets.QTableWidgetItem(key)) # add key vào cột filename
                    self.ui.tableWidget.setItem(row , 1 ,QtWidgets.QTableWidgetItem(self.dictResult[key]))  # add value vào cột path
                    row += 1
            else:
                self.ui.label_error.setText('Không tìm thấy file theo thời gian chỉnh sửa')
        else :   
            for i in range(self.ui.tableWidget.rowCount() -1 ,-1 , -1): # vì xóa theo tuần tự từ row cuối lên row đầu tiên nên cho vòng lặp chạy ngược
                self.ui.tableWidget.removeRow(i)
            self.ui.label_error.setText('Không tìm thấy file bạn muốn')
    def Export(self):
        self.ui.label_error.setText('')
        wb = Workbook()
        ws  = wb.active
        ws.title = 'Exercise4'
        self.nameFileExcel = self.ui.text_export_excel.toPlainText()
        if self.nameFileExcel == '':
             self.ui.label_error.setText('Vui lòng không để trống tên file muốn export')
        else:
            ws.append(['FileName' , 'PathFile']) 
            for i in self.dictResult:
                ws.append([i , self.dictResult[i]])
            wb.save('{}.xlsx'.format(self.nameFileExcel))
            self.ui.label_error.setText('Export thành công') 
    def Import(self):
        try:
            self.ui.label_error.setText('')
            for i in range(self.ui.tableWidget.rowCount() -1 ,-1 , -1): # vì xóa theo tuần tự từ row cuối lên row đầu tiên nên cho vòng lặp chạy ngược
                self.ui.tableWidget.removeRow(i)  #reset lại bảng hiển thị trước khi import
            self.nameFileExcel = self.ui.text_import_excel.toPlainText()
            wb = load_workbook('{}.xlsx'.format(self.nameFileExcel))
            ws = wb.active
            self.dictResult = {} #reset bien dict
            arrFileName = []
            arrPath = []
            for i in range(1 , ws.max_column + 1): 
                for j in range(2,ws.max_row + 1): 
                    if i == 1:
                        arrFileName.append(ws.cell(row = j , column = i).value) 
                    else :
                        arrPath.append(ws.cell(row = j , column = i).value) 
            for i in range(len(arrFileName)):
                self.dictResult[arrFileName[i]] = arrPath[i]              
            row = 0 
            for key in self.dictResult :
                self.ui.tableWidget.insertRow(row) # tạo stt của row
                self.ui.tableWidget.setItem(row , 0,QtWidgets.QTableWidgetItem(key)) # add key vào cột filename
                self.ui.tableWidget.setItem(row , 1 ,QtWidgets.QTableWidgetItem(self.dictResult[key]))  # add value vào cột path
                row += 1
            self.ui.label_error.setText('Import thành công')
        except :
            self.ui.label_error.setText('File không tồn tại')
def main():
    app = QtWidgets.QApplication(sys.argv)
    application = Ecercise5()
    application.show()
    obj = app.exec_()
    sys.exit(obj)

if __name__ == "__main__":
    import sys  
    main()